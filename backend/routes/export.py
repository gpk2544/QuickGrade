"""
Export Route — Excel generation
GET /forums/{forum_id}/export/excel — download .xlsx results
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from routes.auth import verify_token
from firebase_admin import firestore
import io

router = APIRouter()

def db():
    return firestore.client()

@router.get("/{forum_id}/export/excel")
def export_excel(forum_id: str, p=Depends(verify_token)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    # Get forum
    fdoc = db().collection("forums").document(forum_id).get()
    if not fdoc.exists:
        raise HTTPException(404, "Forum not found")
    forum = fdoc.to_dict()

    # Get model answers (for column headers)
    answers = []
    for a in db().collection("model_answers").where("forum_id", "==", forum_id).stream():
        answers.append(a.to_dict())
    # Sort answers by question_num in Python
    answers.sort(key=lambda x: int(x.get("question_num", 0)))

    # Get students
    students = []
    for s in db().collection("students").where("forum_id", "==", forum_id).stream():
        sd = s.to_dict()
        sd["id"] = s.id
        students.append(sd)

    # Sort by reg_number
    students.sort(key=lambda x: x.get("reg_number", ""))

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = forum.get("name", "Results")[:31]

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D3349", end_color="0D3349", fill_type="solid")
    green_fill = PatternFill(start_color="E6F9F2", end_color="E6F9F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")
    red_fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
    summary_fill = PatternFill(start_color="D0F0E8", end_color="D0F0E8", fill_type="solid")
    bold_font = Font(name="Calibri", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    # Title row
    total_marks = forum.get("total_marks", 100)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(answers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{forum.get('name', 'Forum')} — {forum.get('subject', '')} | {forum.get('class_name', '')} | Total: {total_marks}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="0D3349")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Header row
    row = 3
    headers = ["#", "Reg No", "Student Name"]
    for a in answers:
        headers.append(f"Q{a['question_num']} /{a['marks']}")
    headers.extend([f"Total /{total_marks}", "Percentage %", "Result"])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    for idx, s in enumerate(students, 1):
        row += 1
        scores = s.get("scores", {})
        total = s.get("total", 0)
        pct = s.get("percentage", 0)
        passed = pct >= 40

        # Row fill
        if pct >= 75:
            fill = green_fill
        elif pct >= 40:
            fill = yellow_fill
        else:
            fill = red_fill

        values = [idx, s.get("reg_number", ""), s.get("name", "")]
        for a in answers:
            qkey = f"Q{a['question_num']}"
            values.append(scores.get(qkey, 0))
        values.extend([total, f"{pct}%", "PASS" if passed else "FAIL"])

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.fill = fill
            cell.border = thin_border
            if col >= 4:  # Score columns centered
                cell.alignment = center
            # Color result column
            if col == len(values):
                cell.font = Font(name="Calibri", bold=True, color="22C55E" if passed else "EF4444")

    # Summary row
    row += 1
    if students:
        graded = [s for s in students if s.get("status") == "graded"]
        totals = [s.get("total", 0) for s in graded]
        avg = round(sum(totals) / len(totals), 1) if totals else 0
        avg_pct = round((avg / max(total_marks, 1)) * 100, 1) if totals else 0
        passed_count = sum(1 for s in graded if s.get("percentage", 0) >= 40)

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        summary_cell = ws.cell(row=row, column=1, value="CLASS SUMMARY")
        summary_cell.font = bold_font
        summary_cell.fill = summary_fill
        summary_cell.border = thin_border

        for col in range(4, 4 + len(answers)):
            cell = ws.cell(row=row, column=col, value="—")
            cell.fill = summary_fill
            cell.alignment = center
            cell.border = thin_border

        total_col = 4 + len(answers)
        ws.cell(row=row, column=total_col, value=avg).font = bold_font
        ws.cell(row=row, column=total_col).fill = summary_fill
        ws.cell(row=row, column=total_col).alignment = center
        ws.cell(row=row, column=total_col).border = thin_border

        ws.cell(row=row, column=total_col + 1, value=f"{avg_pct}%").font = bold_font
        ws.cell(row=row, column=total_col + 1).fill = summary_fill
        ws.cell(row=row, column=total_col + 1).alignment = center
        ws.cell(row=row, column=total_col + 1).border = thin_border

        ws.cell(row=row, column=total_col + 2, value=f"{passed_count}/{len(graded)} Pass").font = bold_font
        ws.cell(row=row, column=total_col + 2).fill = summary_fill
        ws.cell(row=row, column=total_col + 2).alignment = center
        ws.cell(row=row, column=total_col + 2).border = thin_border

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(3, row + 1))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max(max_len + 3, 10)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = forum.get("name", "Results").replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_Results.xlsx"}
    )
